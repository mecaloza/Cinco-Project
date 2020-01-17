from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.support.ui import Select
from copy import copy, deepcopy

import time
import openpyxl

timeout=20

def get_the_web():
    
    global timeout
    
    # Specifying incognito mode as you launch your browser[OPTIONAL]
    option = webdriver.ChromeOptions()
    option.add_argument("--incognito")
    
    # Create new Instance of Chrome in incognito mode
    browser = webdriver.Chrome('.\drivers\chromedriver', options=option)
    
    # Go to desired website
    browser.get("https://procesos.ramajudicial.gov.co/procesoscs/ConsultaJusticias21.aspx?EntryId=Xsw4o2BqwzV1apD2i2r2orO8yTc%3d")
    # Wait 20 seconds for page to load
    
    
    
    try:
        # Wait until the final element [Avatar link] is loaded.
        # Assumption: If Avatar link is loaded, the whole page would be relatively loaded because it is among
        # the last things to be loaded.
        WebDriverWait(browser, timeout).until(EC.visibility_of_element_located((By.ID, "ddlEntidadEspecialidad")))
    except TimeoutException:
        print("Timed out waiting for page to load")
        browser.quit()
        
    return browser






#------------Get all the elements on the web and send the values on the excel file (database)-------------#

def asignar_nro_proceso ():
    global timeout
    browser=get_the_web()
    
    #Open Excel workbook
    wb_database=openpyxl.load_workbook('Database-Process.xlsx')
    db_sheet=wb_database['Hoja1']

    
    registered_process=len(db_sheet['A'])
    Nproce=1
    
    for i in range(registered_process-1):
        Nproce +=1
    
    
        if(db_sheet.cell(row=Nproce,column=12).value == None):
            
            inputElement2 = browser.find_element_by_id("ddlCiudad")
            inputElement2.send_keys(db_sheet.cell(row=Nproce,column=2).value)
            
            #sleep to give webpage time to load 'Entidades' according to the city name given.
            time.sleep(1)
            
            try:
                WebDriverWait(browser, timeout).until(EC.visibility_of_element_located((By.ID, "ddlEntidadEspecialidad")))
                
                dropdown1 = browser.find_element_by_id("ddlEntidadEspecialidad")
                dropdown1.send_keys(db_sheet.cell(row=Nproce,column=3).value)
                
                #dropdown1= Select(browser.find_element_by_id('ddlEntidadEspecialidad'))
                #dropdown1.select_by_value(db_sheet.cell(row=Nproce,column=3).value)
                
            except TimeoutException:
                print("No se encontro opcion Ciudad")
                browser.quit()
            
            inputElement3 = browser.find_element_by_id("rblConsulta")
            inputElement3.send_keys("Consulta por Nombre o Razón social")
        
            inputElement4 = browser.find_element_by_id("ddlTipoSujeto") 
            inputElement4.send_keys(db_sheet.cell(row=Nproce,column=5).value)
            
            inputElement5 = browser.find_element_by_id("ddlTipoPersona")
            inputElement5.send_keys(db_sheet.cell(row=Nproce,column=7).value)
                
            inputElement6 = browser.find_element_by_id("txtNatural")
            if inputElement6.text != None:
                inputElement6.clear()
            inputElement6.send_keys(db_sheet.cell(row=Nproce,column=4).value)
            
            
            inputElementX=browser.find_element_by_id("sliderBehaviorConsultaNom_railElement")
            inputElementX.click()
    
            inputElement7=browser.find_element_by_id("btnConsultaNom")
            inputElement7.click()
            
            time.sleep(5)
            #get the web element table in which the processes are contained
            tabla_procesos=browser.find_element_by_id('gvResultadosNum')
            #get all the <td> tags of the table in which the data is contained
            campos_tabla_busqueda=tabla_procesos.find_elements_by_tag_name('td')
            
            
    #---------------------------Get all the elements on the web and send the values on the excel file (database)----------------------------#
            
    #---------------------------Assign a process number in the excel file----------------------------#       
    
            #get the number of rows of the table
            cantidad_procesos=len(tabla_procesos.find_elements_by_tag_name('tr'))
    
            lista_numeros_procesos=[]
            lista_fechas_radicacion=[]
                
            #get all the "Fechas de Radicacion", step 7, because dates appear every 7 fields.
            for i in range (2,len(campos_tabla_busqueda),7):
                lista_fechas_radicacion.append(campos_tabla_busqueda[i].text)
            
            #get all the "Numeros de procesos" in the table
            for i in range (cantidad_procesos-1):
            #has to be minus 1 because the heading is included on the list
                try:
                    numero_proceso='gvResultadosNum_lnkActuacionesnNum_'
                    numero_proceso += str(i)   
                    lista_numeros_procesos.append(browser.find_element_by_id(numero_proceso).text)
                except (NoSuchElementException):
                        print('posiblemente hay 2 paginas de procesos')
            
            #get the "fecha radicacion" from the excel file (database) to compare the dates from the table
            fecha_radicacion=db_sheet.cell(row=Nproce,column=9).value
    
            #assign the number process in the excel file
            if fecha_radicacion in lista_fechas_radicacion:
                if lista_fechas_radicacion.count(fecha_radicacion) >1:
                    print('Hay mas de un proceso con la misma fecha, numero no asignado')
                else:
                    db_sheet.cell(row=Nproce,column=12).value= lista_numeros_procesos[lista_fechas_radicacion.index(fecha_radicacion)]
                    wb_database.save('Database-Process.xlsx')
                    print('Numero de Proceso Asignado -  OK')
                    create_excel_file (lista_numeros_procesos[lista_fechas_radicacion.index(fecha_radicacion)])
            else:
                print('Numero de Proceso no asignado')
    print('DONE')
    browser.quit()
         
#---------------------------Assign a process number in the excel file----------------------------#


#time.sleep(2)

def create_excel_file (process_number_given):
    
    browser=get_the_web()
    
    wb_database=openpyxl.load_workbook('Database-Process.xlsx')
    db_sheet=wb_database['Hoja1']
    number_process_column=db_sheet['L']
    process_numbers=[]
    
    for cell in number_process_column:
        process_numbers.append(cell.value)
    
    fila_proceso=(process_numbers.index(process_number_given)+1)


    
    inputElement2 = browser.find_element_by_id("ddlCiudad")
    inputElement2.send_keys(db_sheet.cell(row=fila_proceso,column=2).value)
    time.sleep(1)
    
    dropdown1 = browser.find_element_by_id("ddlEntidadEspecialidad")
    dropdown1.send_keys(db_sheet.cell(row=fila_proceso,column=3).value)
    time.sleep(1)
    
    
    
    inputRadicado = browser.find_element_by_id('divNumRadicacion').find_element_by_tag_name('input')
    inputRadicado.send_keys(process_number_given)
    
    inputElement7=browser.find_element_by_id("sliderBehaviorNumeroProceso_railElement")
    inputElement7.click()
    
    btnconsulta=browser.find_element_by_xpath('/html/body/form/div[2]/table/tbody/tr[2]/td/table/tbody/tr/td/div[3]/table/tbody/tr[4]/td/div[1]/table/tbody/tr[3]/td/input[1]')
    btnconsulta.click()
    
    time.sleep(5)
    
    despacho=browser.find_element_by_id('lblJuzgadoActual').text
    ponente=browser.find_element_by_id('lblPonente').text
    tipo=browser.find_element_by_id('lblTipo').text
    clase=browser.find_element_by_id('lblClase').text
    recurso=browser.find_element_by_id('lblRecurso').text
    ubicacion=browser.find_element_by_id('lblUbicacion').text
    demandantes=browser.find_element_by_id('lblNomDemandante').text
    demandados=browser.find_element_by_id('lblNomDemandado').text
    contenido=browser.find_element_by_id('lblContenido').text
    
    try:
        nombre_documento=browser.find_element_by_id('rptDocumentos_lbNombreDoc_0').text
        descripcion=browser.find_element_by_id('rptDocumentos_lblDescDoc_0').text
    except NoSuchElementException:
        print('No hay documentos Asociados')


    tabla_detalle=browser.find_element_by_class_name('ActuacionesDetalle')
    cantidad_actuaciones=tabla_detalle.find_elements_by_tag_name('tr')
    
    
    lista_fecha_actuaciones=[]
    lista_actuaciones=[]
    lista_anotaciones=[]
    lista_fecha_inicia=[]
    lista_fecha_termina=[]
    lista_fecha_registro=[]



    time.sleep(2)

    #we have to substract 1 , due to cantidad_actuaciones is including the header.
    for i in range(len(cantidad_actuaciones)-1):
    
        fecha_actuacion='rptActuaciones_lblFechaActuacion_'
        actuacion='rptActuaciones_lblActuacion_'
        anotacion='rptActuaciones_lblAnotacion_'
        fecha_inicia='rptActuaciones_lblFechaInicio_'
        fecha_termina='rptActuaciones_lblFechaFin_'
        fecha_registro='rptActuaciones_lblFechaRegistro_'
        
        fecha_actuacion += str(i)
        lista_fecha_actuaciones.append(browser.find_element_by_id(fecha_actuacion).text)
        actuacion += str(i)
        lista_actuaciones.append(browser.find_element_by_id(actuacion).text)
        anotacion += str(i)
        lista_anotaciones.append(browser.find_element_by_id(anotacion).text)
        fecha_inicia += str(i)
        lista_fecha_inicia.append(browser.find_element_by_id(fecha_inicia).text)
        fecha_termina += str(i)
        lista_fecha_termina.append(browser.find_element_by_id(fecha_termina).text)
        fecha_registro += str(i)
        lista_fecha_registro.append(browser.find_element_by_id(fecha_registro).text)
    
    
    #Open Excel Workbook
    
    path='./Procesos/Formato_Base.xlsx'
    excel_name_sheet='Formato'
    wb=openpyxl.load_workbook(path)
    main_sheet=wb[excel_name_sheet]
    
    
    #define the first empty row number, in which the algorythm will write "Actuaciones del Proceso" in Excel
    empty_row=32
    
    while (main_sheet.cell(row = empty_row, column = 3).value != None) :
              empty_row += 1
              print(empty_row)
              
    #fill data in workbook         
              
    main_sheet['C4'].value=despacho
    main_sheet['Z4'].value=ponente
    main_sheet['C9'].value=tipo
    main_sheet['I9'].value=clase
    main_sheet['V9'].value=recurso
    main_sheet['AI9'].value=ubicacion
    main_sheet['C14'].value=demandantes
    main_sheet['Z14'].value=demandados
    main_sheet['C19'].value=contenido
    
    try:
        main_sheet['C26'].value=nombre_documento
        main_sheet['Z26'].value=descripcion
    except UnboundLocalError:
        print('No hay documentos Asociados')
    #Define the cell to copy the style
    
    style_source='C32'
    
    for i in range (len(lista_fecha_actuaciones)):
        
        
        main_sheet.cell(row=(empty_row+i),column=3).value=lista_fecha_actuaciones[i]
        main_sheet.cell(row=(empty_row+i),column=7).value=lista_actuaciones[i]
        main_sheet.cell(row=(empty_row+i),column=11).value=lista_anotaciones[i]
        main_sheet.cell(row=(empty_row+i),column=37).value=lista_fecha_inicia[i]
        main_sheet.cell(row=(empty_row+i),column=41).value=lista_fecha_termina[i]
        main_sheet.cell(row=(empty_row+i),column=45).value=lista_fecha_registro[i]
        main_sheet.row_dimensions[empty_row+i].height = 33
        
        main_sheet.cell(row=(empty_row+i), column=3)._style=deepcopy(main_sheet[style_source]._style)
        main_sheet.merge_cells(start_row=empty_row+i, start_column=3, end_row=empty_row+i, end_column=3+3)
        main_sheet.cell(row=(empty_row+i), column=7)._style=deepcopy(main_sheet[style_source]._style)
        main_sheet.merge_cells(start_row=empty_row+i, start_column=7, end_row=empty_row+i, end_column=7+3)
        main_sheet.cell(row=(empty_row+i), column=11)._style=deepcopy(main_sheet[style_source]._style)
        main_sheet.merge_cells(start_row=empty_row+i, start_column=11, end_row=empty_row+i, end_column=11+25)
        main_sheet.cell(row=(empty_row+i), column=37)._style=deepcopy(main_sheet[style_source]._style)
        main_sheet.merge_cells(start_row=empty_row+i, start_column=37, end_row=empty_row+i, end_column=37+3)
        main_sheet.cell(row=(empty_row+i), column=41)._style=deepcopy(main_sheet[style_source]._style)
        main_sheet.merge_cells(start_row=empty_row+i, start_column=41, end_row=empty_row+i, end_column=41+3)
        main_sheet.cell(row=(empty_row+i), column=45)._style=deepcopy(main_sheet[style_source]._style)
        main_sheet.merge_cells(start_row=empty_row+i, start_column=45, end_row=empty_row+i, end_column=45+3)

    browser.quit()
    new_path="./Procesos/" + process_number_given + '.xlsx'
    wb.save(new_path)
 
    
    print('Excel creado exitosamente, Cesar Crack')

