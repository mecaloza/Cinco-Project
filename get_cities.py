# -*- coding: utf-8 -*-
"""
Created on Wed Jan 15 21:33:25 2020

@author: Cesar
"""
from webscraping import get_the_web
import time
import openpyxl


# Este proceso debe hacerse dentro de los chequeos periodicos para garantizar que no se hayan añadido nuevas ciudades o destinos
def get_cities_entities_web():
    
    wb_cities_entities=openpyxl.load_workbook('Cities-Entities.xlsx')
    cities_entities_sheet=wb_cities_entities['Cities-Entities']
    row_cell=1
    
    browser=get_the_web()
    obj_ciudad=(browser.find_element_by_id('ddlCiudad')).find_elements_by_tag_name('option')
    ciudades=[]
    
    
    for i in obj_ciudad:
        if i.get_attribute('value') != '0':
            ciudades.append(i.text)

    
    lista_prueba=[]
    for i in range(len(ciudades)):
        browser.find_element_by_id('ddlCiudad').send_keys(ciudades[i])
        time.sleep(2)
        obj_entidad=(browser.find_element_by_id('ddlEntidadEspecialidad')).find_elements_by_tag_name('option')

        for j in range(len(obj_entidad)):
            if obj_entidad[j].get_attribute('value') !='0':
                row_cell += 1
                cities_entities_sheet.cell(row=row_cell,column=1).value=ciudades[i]
                lista_prueba.append(obj_entidad[j].text)
                cities_entities_sheet.cell(row=row_cell,column=2).value=obj_entidad[j].text
            
    wb_cities_entities.save('Cities-Entities.xlsx')
    
    
    print('done')
    
def make_cities_entities_dictionary():

    wb_cities_entities=openpyxl.load_workbook('Cities-Entities.xlsx')
    cities_entities_sheet=wb_cities_entities['Cities-Entities']
    cantidad_celdas=(len(cities_entities_sheet['A']))
    
    diccionario={}
    
    ciudades_choices=[]
    entidades_choices=[]
    
    for cell in cities_entities_sheet['A']:
        ciudades_choices.append(cell.value)
        
    for cell in cities_entities_sheet['B']:
        entidades_choices.append(cell.value)
    
    entidades_ciudad=[]
    ciudades_no_repetidas= sorted(set(ciudades_choices))

    for i in range(2,cantidad_celdas):
        
        if i == (cantidad_celdas-1):
            entidades_ciudad.append(entidades_choices[i])
            diccionario[ciudades_choices[i]]=entidades_ciudad
            break
        
        elif ciudades_choices[i]==ciudades_choices[i+1]:
            entidades_ciudad.append(entidades_choices[i])
    
        else:
             entidades_ciudad.append(entidades_choices[i])
             diccionario[ciudades_choices[i]]=entidades_ciudad
             entidades_ciudad=[]
    
    return diccionario,ciudades_no_repetidas

