# -*- coding: utf-8 -*-
"""
Created on Thu Nov 28 14:28:41 2019

@author: user
"""
from selenium import webdriver
import wx
import time
import openpyxl
from webscraping import asignar_nro_proceso, get_the_web
from get_cities import get_cities_entities_web, make_cities_entities_dictionary
import os


DB = openpyxl.load_workbook('Database-Process.xlsx')
sheet = DB['Hoja1']


class MyFrame(wx.Frame):
    
    
    def OnKeyDown(self, event):
        """quit if user press q or Esc"""
        if event.GetKeyCode() == 27 or event.GetKeyCode() == ord('Q'): #27 is Esc
            self.Close(force=True)
            
        else:
            event.Skip()
 
    def __init__(self):
        ancho,alto=wx.DisplaySize()
        wx.Frame.__init__(self, None, wx.ID_ANY, "Software Legal", size=(1200, 700))  
        self.Bind(wx.EVT_KEY_UP, self.OnKeyDown)
        
        try:
            image_file = 'CINCO CONSULTORES.jpg'
            bmp1 = wx.Image(image_file, wx.BITMAP_TYPE_ANY).ConvertToBitmap()
            
            self.panel = wx.StaticBitmap(self, -1, bmp1, (0, 0))
            
        except IOError:
            print ("Image file %s not found"  )
            raise SystemExit
            
            
        button = wx.Button(self.panel, id=wx.ID_ANY, label="Ingresar Proceso" ,pos=(900, 100), size=(200, 50))
        button.Bind(wx.EVT_BUTTON, self.Ingresarproceso)
        
        button2 = wx.Button(self.panel, id=wx.ID_ANY, label="Consultar Proceso" ,pos=(900, 150), size=(200, 50))
        button2.Bind(wx.EVT_BUTTON, self.BtnConsultaProceso)
        
        button3 = wx.Button(self.panel, id=wx.ID_ANY, label="Actualizar información proceso" ,pos=(900, 200), size=(200, 50))
        button3.Bind(wx.EVT_BUTTON, self.onButton3)
        
        btn_asignar_procesos = wx.Button(self.panel, id=wx.ID_ANY, label="Ident. Nro Proceso" ,pos=(900, 250), size=(200, 50))
        btn_asignar_procesos.Bind(wx.EVT_BUTTON, self.onBtn_asignar_procesos)
        
        ico = wx.Icon('Icono.ico', wx.BITMAP_TYPE_ICO)
        self.SetIcon(ico)

    
    #-------------Button Functions-----------------#
    def Ingresarproceso(self, event):
        secondWindow = ww_Ingresar_Proceso(parent=self.panel)
        secondWindow.Show()

    def BtnConsultaProceso(self, event): 
        consultawindow=ww_Consultar_Proceso(parent=self.panel)
        consultawindow.Show()

        
    def onButton3(self, event):
        print ("Button pressed!")
        
    def onBtn_asignar_procesos(self, event):
        asignar_nro_proceso()
        
    #-------------Button Functions-----------------#    

        
class ww_Ingresar_Proceso(wx.Frame):
   
    ciudades_entidades=make_cities_entities_dictionary()
    title = "Ingresar Proceso"
    
    def __init__(self,parent):
        ancho,alto=wx.DisplaySize()
        ciudades_entidades=make_cities_entities_dictionary()
        wx.Frame.__init__(self,parent, -1,'Ingresar Proceso', size=(1300,700))   

        try:
            
            image_file = 'CINCO CONSULTORES.jpg'
            bmp1 = wx.Image(
                image_file, 
                wx.BITMAP_TYPE_ANY).ConvertToBitmap()
            
            self.panel = wx.StaticBitmap(
                self, -1, bmp1, (0, 0))
            self.panel.SetBackgroundColour(wx.Colour('white'))
            
            
        except IOError:
            print ("Image file %s not found"  )
            raise SystemExit
        
        
        ico = wx.Icon('Icono.ico', wx.BITMAP_TYPE_ICO)
        title_font= wx.Font(25, wx.FONTFAMILY_DECORATIVE, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD)
        
        self.SetIcon(ico)
        
        self.lbltitle =wx.StaticText(self.panel, label='Nuevo Proceso')
        self.lbltitle.SetFont(title_font)
        self.lbltitle.SetBackgroundColour('white')
        
        self.lblciudad = wx.StaticText(self.panel, label="Ciudad:", pos=(600, 50))
        self.lblciudad.SetBackgroundColour("white")
        self.Ciudad=wx.ComboBox(self.panel, choices=ciudades_entidades[1],id=wx.ID_ANY, pos=(700,45),size=(100, -1))
        #self.Ciudad = wx.TextCtrl(self.panel, size=(100, -1),pos=(700, 45))
        self.Ciudad.Bind(wx.EVT_COMBOBOX, self.get_entidades)
    
        self.lblentidad = wx.StaticText(self.panel, label="Entidad:", pos=(600, 100))
        self.lblentidad.SetBackgroundColour("white")
        self.Entidades=wx.ComboBox(self.panel, choices=ciudades_entidades[0]['MEDELLIN '], pos=(700,95),size=(100, -1))      

        self.lbljurisdiccion = wx.StaticText(self.panel, label="Jurisdicción:",pos=(600, 150))
        self.lbljurisdiccion.SetBackgroundColour("white")
        self.Jurisdi = wx.TextCtrl(self.panel, size=(100, -1),pos=(700, 145))
        
        self.lblraz_soc = wx.StaticText(self.panel, label="Por Nombre o \npor razon Social:",pos=(600, 250))
        self.lblraz_soc.SetBackgroundColour("white")
        self.razon = wx.TextCtrl(self.panel, size=(100, -1),pos=(700, 245))
        
        self.lbltipo_sujeto = wx.StaticText(self.panel, label="Tipo Sujeto:",pos=(600, 350))
        self.lbltipo_sujeto.SetBackgroundColour("white")
        self.Tipsuj = wx.ComboBox(self.panel, choices=['Demandado','Demandante'], size=(100, -1),pos=(700, 345))
        
        self.lblresponsable = wx.StaticText(self.panel, label="Responsable:",pos=(600, 450))
        self.lblresponsable.SetBackgroundColour("white")
        self.Responsable = wx.TextCtrl(self.panel, size=(100, -1),pos=(700, 445))
        
        self.lbltipo_persona = wx.StaticText(self.panel, label="Tipo de Persona:",pos=(600, 550))
        self.lbltipo_persona.SetBackgroundColour("white")
        self.Tipopersona = wx.ComboBox(self.panel, choices=['Natural','Juridica'], size=(100, -1),pos=(700, 545))
        
        self.lblnombre = wx.StaticText(self.panel, label="Nombre:",pos=(900, 50))
        self.lblnombre.SetBackgroundColour("white")
        self.Nombre= wx.TextCtrl(self.panel, size=(100, -1),pos=(1050, 45))
        
        self.lblfecha_rad = wx.StaticText(self.panel, label="Fecha de Radicación:",pos=(900, 150))
        self.lblfecha_rad.SetBackgroundColour("white")
        self.Fechara = wx.TextCtrl(self.panel, size=(100, -1),pos=(1050, 145))
        
        self.lblcedula = wx.StaticText(self.panel, label="Cedula:",pos=(900, 250))
        self.lblcedula.SetBackgroundColour("white")
        self.Cedula = wx.TextCtrl(self.panel, size=(100, -1),pos=(1050, 245))
        
        self.lblapoderado = wx.StaticText(self.panel, label="Apoderado:",pos=(900, 350))
        self.lblapoderado.SetBackgroundColour("white")
        self.Apoderado = wx.TextCtrl(self.panel, size=(100, -1),pos=(1050, 345))
        
        button = wx.Button(self.panel, id=wx.ID_ANY, label="Crear Proceso" ,pos=(900, 400), size=(200, 100))
        button.Bind(wx.EVT_BUTTON, self.Crearproceso)
        
        button = wx.Button(self.panel, id=wx.ID_ANY, label="Cancelar" ,pos=(900, 550), size=(200, 100))
        button.Bind(wx.EVT_BUTTON, self.OnCloseWindow)
        
        self.SetBackgroundColour(wx.Colour(100,100,100))
        self.Centre(True)
        self.Show()
        
        mainSizer= wx.BoxSizer(wx.VERTICAL)
        hSizer=wx.BoxSizer(wx.HORIZONTAL)
        
        line1sizer=wx.BoxSizer()
        line2sizer=wx.BoxSizer()
        line3sizer=wx.BoxSizer()
        line4sizer=wx.BoxSizer()
        line5sizer=wx.BoxSizer()
        line6sizer=wx.BoxSizer()
        line7sizer=wx.BoxSizer()
        line8sizer=wx.BoxSizer()
        line9sizer=wx.BoxSizer()
        line10sizer=wx.BoxSizer()
        line11sizer=wx.BoxSizer()

        hSizer.Add(self.lbltitle, proportion=0, flag=wx.ALL, border=5)
        line1sizer.Add(self.lblciudad, proportion=0,flag=wx.ALL,border=5)
        line1sizer.Add(self.Ciudad, proportion=0,flag=wx.ALL,border=5) 
        line2sizer.Add(self.lblentidad, proportion=0,flag=wx.ALL,border=5)
        line2sizer.Add(self.Entidades, proportion=0,flag=wx.ALL,border=5)
        line3sizer.Add(self.lbljurisdiccion, proportion=0,flag=wx.ALL,border=5)
        line3sizer.Add(self.Jurisdi, proportion=0,flag=wx.ALL,border=5)        
        line4sizer.Add(self.lbltipo_sujeto, proportion=0,flag=wx.ALL,border=5)
        line4sizer.Add(self.Tipsuj, proportion=0,flag=wx.ALL,border=5)
        #line5sizer.Add(self.lblname10, proportion=0,flag=wx.ALL,border=5)
        line6sizer.Add(self.lbltipo_persona, proportion=0,flag=wx.ALL,border=5)
        line6sizer.Add(self.Tipopersona, proportion=0,flag=wx.ALL,border=5)
        line7sizer.Add(self.lblraz_soc, proportion=0,flag=wx.ALL,border=5)
        line7sizer.Add(self.razon, proportion=0,flag=wx.ALL,border=5)
        #line8sizer.Add(self.lbl, proportion=0,flag=wx.ALL,border=5)
        #line9sizer.Add(self.lbltipo, proportion=0,flag=wx.ALL,border=5)
        #line10sizer.Add(self.lblname10, proportion=0,flag=wx.ALL,border=5)
        #line11sizer.Add(self.lblname10, proportion=0,flag=wx.ALL,border=5)
        
        mainSizer.Add(hSizer,0, flag=wx.ALIGN_CENTER)
        mainSizer.Add(line1sizer,0, flag=wx.LEFT)
        mainSizer.Add(line2sizer,0, flag=wx.LEFT)
        mainSizer.Add(line3sizer,0, flag=wx.LEFT)
        mainSizer.Add(line4sizer,0, flag=wx.LEFT)
        mainSizer.Add(line5sizer,0, flag=wx.LEFT)
        mainSizer.Add(line6sizer,0, flag=wx.LEFT)
        mainSizer.Add(line7sizer,0, flag=wx.LEFT)
        mainSizer.Add(line8sizer,0, flag=wx.LEFT)
        mainSizer.Add(line9sizer,0, flag=wx.LEFT)
        mainSizer.Add(line10sizer,0, flag=wx.LEFT)
        mainSizer.Add(line11sizer,0, flag=wx.LEFT)
        
        
 
        
        self.SetSizer(mainSizer)
        #mainSizer.Fit(self)
        self.Layout()
        
        
        
    def OnCloseWindow(self, event):
        self.Destroy()

    def get_entidades(self):
        global Ciudad
        Ciudad = self.Ciudad.GetValue()
        return Ciudad
        
    def Crearproceso(self, event):
        
        Nproce = 1
        
        while (sheet.cell(row = Nproce, column = 1).value != None) :
          Nproce = Nproce + 1
          print(Nproce)

        Ciudad = self.Ciudad.GetValue()
        sheet.cell(row = Nproce, column = 2).value = Ciudad
        
        Jurisdi = self.Jurisdi.GetValue()
        sheet.cell(row = Nproce, column = 3).value = Jurisdi
        
        Entidad = self.Entidades.GetValue()
        sheet.cell(row = Nproce, column = 3).value = Entidad
        
        razon= self.razon.GetValue()
        sheet.cell(row = Nproce, column = 4).value = razon
        
        Tipsuj = self.Tipsuj.GetValue()
        sheet.cell(row = Nproce, column = 5).value = Tipsuj 
        
        Responsable = self.Responsable.GetValue()
        sheet.cell(row = Nproce, column = 6).value = Responsable
        
        Tipopersona = self.Tipopersona.GetValue()
        sheet.cell(row = Nproce, column = 7).value = Tipopersona
        
        Nombre  = self.Nombre.GetValue()
        sheet.cell(row = Nproce, column = 8).value = Nombre
        
        Fechara  = self.Fechara.GetValue()
        sheet.cell(row = Nproce, column = 9).value = Fechara
        
        Cedula= self.Cedula.GetValue()
        sheet.cell(row = Nproce, column = 10).value = Cedula
        
        Apoderado = self.Apoderado.GetValue()
        sheet.cell(row = Nproce, column = 11).value =Apoderado
         
        sheet.cell(row = Nproce  , column = 1).value = Nproce 
        
        
#        filepath = "C:\Users\user\.spyder-py3\" 

        DB.save('Database-Process.xlsx')

class ww_Consultar_Proceso(wx.Frame):
    

    def __init__(self,parent):
        wx.Frame.__init__(self,parent, -1,'Consultar Proceso', size=(1200,700))   

        try:
            
            image_file = 'CINCO CONSULTORES.jpg'
            bmp1 = wx.Image(
                image_file, 
                wx.BITMAP_TYPE_ANY).ConvertToBitmap()
            
            self.panel = wx.StaticBitmap(
                self, -1, bmp1, (0, 0))
           
       
          
        except IOError:
            print ("Image file %s not found"  )
            raise SystemExit
        
        
        ico = wx.Icon('Icono.ico', wx.BITMAP_TYPE_ICO)
        self.SetIcon(ico)
        
        self.lblname1 = wx.StaticText(self.panel, label="Ingrese Numero de Proceso", pos=(830, 250))
        self.lblname1.SetBackgroundColour("white")
        self.numero_consulta=wx.TextCtrl(self.panel, size=(300, -1),pos=(750, 270))

        btn_consultar = wx.Button(self.panel, id=wx.ID_ANY, label="Consultar" ,pos=(840, 300), size=(100, 30))
        btn_consultar.Bind(wx.EVT_BUTTON, self.Consultar_Excel)
        
    def Consultar_Excel(self,event):
        
        numero_consulta=self.numero_consulta.GetValue()
        print(os.getcwd())
        workbook_path=os.getcwd()+'/Procesos/'+ numero_consulta + '.xlsx'
        os.startfile(workbook_path)
        
class MyApp(wx.App):
    def OnInit(self):
        self.frame= MyFrame()
        self.frame.Show()
        return True       
 
# Run the program     
app=MyApp()
app.MainLoop()
del app